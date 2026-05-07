import os
from datetime import date, datetime
from typing import List
from ..models.task import Task
from ..utils.enums import TaskStatus, Priority


def _find_chinese_font() -> str:
    """查找系统中可用的中文字体路径。"""
    candidates = [
        "C:/Windows/Fonts/msyh.ttc",      # 微软雅黑
        "C:/Windows/Fonts/msyhbd.ttc",     # 微软雅黑粗体
        "C:/Windows/Fonts/simhei.ttf",     # 黑体
        "C:/Windows/Fonts/simsun.ttc",     # 宋体
        "C:/Windows/Fonts/simkai.ttf",     # 楷体
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return ""


def export_xlsx(tasks: List[Task], file_path: str):
    """将任务树导出为 xlsx 文件，包含层级信息。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "任务列表"

    headers = ["层级", "任务名称", "状态", "进度(%)", "优先级", "开始日期", "截止日期", "风险", "备注"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    priority_fills = {
        1: PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid"),
        2: PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        3: PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        4: PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        5: PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    }

    row_idx = 2

    def write_row(task: Task, level: int):
        nonlocal row_idx
        level_prefix = "  " * level + ("└ " if level > 0 else "")
        values = [
            level,
            level_prefix + task.name,
            task.status,
            task.progress,
            Priority(task.priority).label if 1 <= task.priority <= 5 else "中",
            task.start_date.isoformat() if task.start_date else "",
            task.due_date.isoformat() if task.due_date else "",
            task.risk,
            task.remarks,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col == 5:
                cell.fill = priority_fills.get(task.priority, priority_fills[3])
        row_idx += 1
        for child in task.children:
            write_row(child, level + 1)

    for t in tasks:
        write_row(t, 0)

    col_widths = [8, 30, 10, 10, 10, 14, 14, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(file_path)


def import_xlsx(file_path: str) -> List[Task]:
    """从 xlsx 文件导入任务，还原层级关系。"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    ws = wb.active
    tasks = []
    stack = [(-1, None)]

    priority_label_map = {"极低": 1, "低": 2, "中": 3, "高": 4, "紧急": 5}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        level = int(row[0]) if row[0] is not None else 0
        name = str(row[1]).strip().lstrip("└ ")
        status = str(row[2]) if row[2] else TaskStatus.TODO.value
        progress = int(row[3]) if row[3] is not None else 0
        priority_label = str(row[4]) if row[4] else "中"
        priority = priority_label_map.get(priority_label, 3)
        start_date = date.fromisoformat(str(row[5])) if row[5] else None
        due_date = date.fromisoformat(str(row[6])) if row[6] else None
        risk = str(row[7]) if row[7] else ""
        remarks = str(row[8]) if row[8] else ""

        while stack and stack[-1][0] >= level:
            stack.pop()

        parent = stack[-1][1] if stack else None
        parent_id = parent.id if parent else None

        task = Task(
            parent_id=parent_id, name=name, progress=progress,
            status=status, priority=priority, start_date=start_date,
            due_date=due_date, risk=risk, remarks=remarks,
        )
        if parent:
            parent.children.append(task)
        else:
            tasks.append(task)

        stack.append((level, task))

    return tasks


def export_pdf(tasks: List[Task], file_path: str):
    """将任务树导出为 PDF 文件（支持中文）。"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise ImportError("导出 PDF 需要安装 reportlab: pip install reportlab")

    # 注册中文字体
    font_path = _find_chinese_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
            chinese_font = 'ChineseFont'
        except Exception:
            chinese_font = 'Helvetica'
    else:
        chinese_font = 'Helvetica'

    doc = SimpleDocTemplate(file_path, pagesize=A4)
    styles = getSampleStyleSheet()

    # 覆盖默认字体为中文字体
    for style_name in styles.byName:
        s = styles[style_name]
        s.fontName = chinese_font

    title_style = styles["Title"]
    title_style.fontName = chinese_font
    title_style.fontSize = 18

    elements = [Paragraph("AuraTasks Pro - 任务列表", title_style), Spacer(1, 10 * mm)]

    headers = ["任务名称", "状态", "进度", "优先级", "截止日期", "风险"]
    data = [headers]

    def add_rows(task: Task, level: int):
        prefix = "  " * level + ("└ " if level > 0 else "")
        data.append([
            prefix + task.name,
            task.status,
            f"{task.progress}%",
            Priority(task.priority).label if 1 <= task.priority <= 5 else "中",
            task.due_date.isoformat() if task.due_date else "",
            task.risk,
        ])
        for child in task.children:
            add_rows(child, level + 1)

    for t in tasks:
        add_rows(t, 0)

    col_widths = [55 * mm, 18 * mm, 15 * mm, 18 * mm, 25 * mm, 55 * mm]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), chinese_font),
        ('BACKGROUND', (0, 0), (-1, 0), HexColor("#1E1E1E")),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor("#555555")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor("#2C2C2E"), HexColor("#3A3A3C")]),
        ('TEXTCOLOR', (0, 1), (-1, -1), HexColor("#F5F5F7")),
    ]))
    elements.append(table)
    doc.build(elements)
